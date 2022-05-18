import json
import os.path
import openpyxl
import re
import shutil
from openpyxl import load_workbook
from functions import *

print('Программа запущена!!!')

fileNames = []

rootDir = os.getcwd()
dataDir = os.path.join(rootDir, 'data')
resultDir = os.path.join(rootDir, 'result')

print('Очищаем папку со старыми результатами!')
if not (os.path.isdir(resultDir)):
    os.mkdir(resultDir)
else:
    shutil.rmtree(resultDir)
    os.mkdir(resultDir)

configFile = os.path.join(rootDir, 'config.txt')
with open(configFile, encoding='windows-1251') as file:
    configs = json.load(file)

fileNameConfig = configs['fileName']
if (len(fileNameConfig) > 0):
    filePath = os.path.join(dataDir, fileNameConfig)
    if (os.path.exists(filePath)):
        fileNames.append(fileNameConfig)
    else:
        print('Ошибка!!! Файл {} не существует'.format(fileNameConfig))
else:
    filesRaw = os.listdir(dataDir)
    for fileRaw in filesRaw:
        isXlsx = fileRaw.lower().endswith('.xlsx')
        if (isXlsx == True):
            fileNames.append(fileRaw)

if not (len(fileNames) > 0):
    print('Ошибка!!! Не найдено ни одного файла!')
    exit()

for fileName in fileNames:
    print('Работаем с файлом {}'.format(fileName))
    countFiles = int(configs['countFiles'])
    countFilesMass = re.findall('_{2}([0-9]+)_{2}', fileName)
    if (len(countFilesMass) > 0):
        countFiles = int(countFilesMass[0])

    fileFrom = os.path.join(dataDir, fileName)
    fileNameMass = fileName.split('.')
    resultDirNew = os.path.join(resultDir, fileNameMass[0])
    if not (os.path.isdir(resultDirNew)):
        os.mkdir(resultDirNew)

    wb = load_workbook(filename=fileFrom)
    sheet_first = wb.sheetnames[0]
    sheet = wb[sheet_first]
    ws_max_row = sheet.max_row
    ws_max_col = sheet.max_column

    rows_float = ws_max_row / countFiles
    rows_mod = ws_max_row // countFiles
    rows_on_file = 0
    if rows_float > rows_mod:
        rows_on_file = rows_mod + 1
    else:
        rows_on_file = rows_mod

    file_num = 0
    rows = []
    for r in range(1, ws_max_row + 1):
        row = []
        for c in range(1, ws_max_col + 1):
            row.append(sheet.cell(row=r, column=c).value)
        rows.append(row)

        if(len(rows) >= rows_on_file):
            file_num += 1
            saveToFile(resultDirNew, fileName, file_num, rows)
            print('Для файла {} создаётся {} файл'.format(fileName, file_num))
            rows = []

    if(len(rows) > 0):
        file_num += 1
        saveToFile(resultDirNew, fileName, file_num, rows)
        print('Для файла {} создаётся {} файл'.format(fileName, file_num))
        rows = []

print('Программа выполнена!!!')
input('Для выхода из программы нажмите Enter!')
exit()