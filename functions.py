import os
import openpyxl

def saveToFile(dir, fileNameRow, file_num, rows):
    fileMass = fileNameRow.split('.')
    del fileMass[-1]
    fileName = ''.join(fileMass) + '_' + str(file_num) + '.xlsx'
    filePath = os.path.join(dir, fileName)

    # Создаём файл excel
    wb = openpyxl.Workbook()
    ws = wb.create_sheet('Лист1', 0)

    num_row = 0
    num_col = 0
    for row_1 in rows:
        num_row += 1
        for col_1 in row_1:
            num_col += 1
            ws.cell(row=num_row, column=num_col).value = col_1
        num_col = 0

    wb.save(filePath)
    wb.close()